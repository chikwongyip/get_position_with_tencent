"""
@Description : 批量获取xlsx中的位置经纬度
@File        : main.py
@Project     : txmap
@Time        : 2020/4/4 17:55
@Author      : Dexter
@Software    : PyCharm
"""
import os

import openpyxl
import requests 
from dotenv import load_dotenv

load_dotenv()

KEY = 'N7HBZ-LYELZ-DSUXW-TCFZT-QURCZ-B6FZU'

XLSX_PATH = os.path.join(os.getcwd(), '位置.xlsx')



def get_cor(address):
    """
    获取单个地址对应坐标
    :param address: 地址名称
    :return: 字典类型的经纬度坐标,格式:{'lat': 36.112438, 'lng': 114.393631}
    """
    
    ret = requests.get(
        url='https://apis.map.qq.com/ws/geocoder/v1/',
        params={
            'address': address,
            'key': KEY
        }
    ).json()
    print('ret',ret)
    if ret.get('status') == 0:
        return ret.get('result').get('location')
    else:
        return None


if __name__ == '__main__':
  
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['Sheet1']
   
    for cell in ws['A']:
        # 第一行是列标题,需要排除掉
        if cell.row > 1:
            addr = cell.value
            print('addr',addr)
            cor = get_cor(addr)
            print('cor',cor)
            if cor:
                # 经度
                ws[f'B{cell.row}'] = cor.get('lng')
                # 维度
                ws[f'C{cell.row}'] = cor.get('lat')

    wb.save(XLSX_PATH)
